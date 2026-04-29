"""
Organize runs_gm experiment results into categorized CSV/Excel files.

Categories:
  1. single_scene: single_scene_only==True or num_scenes in {0,1} or "scene0" in name
  2. 10scene: num_scenes == 10
  3. 90scene: num_scenes == -1 (all scenes)

Only includes runs that have an 'inference' subfolder (valid runs).
Records detailed args and PSNR performance.
"""

import os
import json
import csv
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUNS_DIR = os.path.join(BASE_DIR, "runs_gm")
OUTPUT_DIR = os.path.join(BASE_DIR, "runs_gm_summary")


def load_args(run_path):
    """Load args.json from a run directory."""
    args_path = os.path.join(run_path, "args.json")
    if os.path.exists(args_path):
        with open(args_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def get_latest_epoch_from_logs(run_path):
    """Parse tensorboard logs to find the maximum logged step (epoch)."""
    logs_dir = os.path.join(run_path, "logs")
    if not os.path.isdir(logs_dir):
        return None
    try:
        from tensorboard.backend.event_processing.event_accumulator import EventAccumulator
        # Only load scalars to save memory and time
        ea = EventAccumulator(logs_dir, size_guidance={'scalars': 0})
        ea.Reload()
        tags = ea.Tags().get("scalars", [])
        max_epoch = -1
        for tag in tags:
            try:
                events = ea.Scalars(tag)
                if events:
                    max_step = max(e.step for e in events)
                    if max_step > max_epoch:
                        max_epoch = max_step
            except Exception:
                pass
        return max_epoch if max_epoch >= 0 else None
    except ImportError:
        # print("    [NOTE] tensorboard package missing, skipping log parse.")
        return None
    except Exception:
        return None


def extract_psnr_info(run_path):
    """
    Extract PSNR info from:
      1. psnr_best_psnr.json  (preferred - has epoch + per-plane PSNR)
      2. psnr_best.json       (fallback)
      3. inference/summary.json (has avg_psnr)
    Returns dict with: avg_psnr, best_epoch, per_plane_psnr (list), raw_data
    """
    result = {
        "avg_psnr": None,
        "best_epoch": None,
        "latest_epoch": None,
        "plane0_psnr": None,
        "plane20_psnr": None,
        "plane39_psnr": None,
        "ckpt_tag": None,
    }

    # --- 1) inference/best_psnr/results.json (Preferred) ---
    results_path = os.path.join(run_path, "inference", "best_psnr", "results.json")
    if os.path.exists(results_path):
        try:
            with open(results_path, "r", encoding="utf-8") as f:
                planes = json.load(f)
            psnr_vals = []
            for p in planes:
                plane_idx = p.get("plane_idx", p.get("plane"))
                psnr_val = p.get("psnr")
                if psnr_val is not None:
                    psnr_vals.append(psnr_val)
                if plane_idx == 0:
                    result["plane0_psnr"] = psnr_val
                elif plane_idx == 20:
                    result["plane20_psnr"] = psnr_val
                elif plane_idx == 39:
                    result["plane39_psnr"] = psnr_val
            if psnr_vals:
                result["avg_psnr"] = sum(psnr_vals) / len(psnr_vals)
        except Exception as e:
            print(f"    [WARN] Error reading {results_path}: {e}")

    # --- 2) Fallback: psnr_best_psnr.json or psnr_best.json ---
    if result["avg_psnr"] is None:
        for psnr_file in ["psnr_best_psnr.json", "psnr_best.json"]:
            path = os.path.join(run_path, psnr_file)
            if os.path.exists(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    result["best_epoch"] = data.get("epoch")
                    result["ckpt_tag"] = data.get("tag")
                    planes = data.get("results", [])
                    psnr_vals = []
                    for p in planes:
                        plane_idx = p.get("plane_idx", p.get("plane"))
                        psnr_val = p.get("psnr")
                        if psnr_val is not None:
                            psnr_vals.append(psnr_val)
                        if plane_idx == 0:
                            result["plane0_psnr"] = psnr_val
                        elif plane_idx == 20:
                            result["plane20_psnr"] = psnr_val
                        elif plane_idx == 39:
                            result["plane39_psnr"] = psnr_val
                    if psnr_vals:
                        result["avg_psnr"] = sum(psnr_vals) / len(psnr_vals)
                    break  # Use first found
                except Exception as e:
                    print(f"    [WARN] Error reading {path}: {e}")

    # --- 2) inference/summary.json for avg_psnr ---
    summary_path = os.path.join(run_path, "inference", "summary.json")
    if os.path.exists(summary_path):
        try:
            with open(summary_path, "r", encoding="utf-8") as f:
                summary = json.load(f)
            # summary can have keys like "best_psnr", "best", "latest"
            for tag_key in ["best_psnr", "best", "latest"]:
                if tag_key in summary:
                    tag_data = summary[tag_key]
                    if "avg_psnr" in tag_data:
                        if result["avg_psnr"] is None:
                            result["avg_psnr"] = tag_data["avg_psnr"]
                        # Use summary avg_psnr if we already have one from psnr file, prefer summary
                        result[f"summary_{tag_key}_avg_psnr"] = tag_data["avg_psnr"]
                    # Also extract per-plane from summary if not yet set
                    if "results" in tag_data and result["plane0_psnr"] is None:
                        for p in tag_data["results"]:
                            plane_idx = p.get("plane_idx", p.get("plane"))
                            psnr_val = p.get("psnr")
                            if plane_idx == 0 and result["plane0_psnr"] is None:
                                result["plane0_psnr"] = psnr_val
                            elif plane_idx == 20 and result["plane20_psnr"] is None:
                                result["plane20_psnr"] = psnr_val
                            elif plane_idx == 39 and result["plane39_psnr"] is None:
                                result["plane39_psnr"] = psnr_val
        except Exception as e:
            print(f"    [WARN] Error reading summary: {e}")

    # --- 3) Check inference subdirs for infer_config.json to get epoch ---
    if result["best_epoch"] is None:
        infer_config_path = os.path.join(run_path, "inference", "best_psnr", "infer_config.json")
        if os.path.exists(infer_config_path):
            try:
                with open(infer_config_path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                result["best_epoch"] = cfg.get("epoch")
            except Exception:
                pass

    # --- 4) Parse tensorboard logs directly for the max epoch (Primary) ---
    log_epoch = get_latest_epoch_from_logs(run_path)
    if log_epoch is not None:
        result["latest_epoch"] = log_epoch

    # --- 5) Fallback: Check psnr_latest.json or inference/latest/infer_config.json ---
    if result["latest_epoch"] is None:
        latest_path = os.path.join(run_path, "psnr_latest.json")
        if os.path.exists(latest_path):
            try:
                with open(latest_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                result["latest_epoch"] = data.get("epoch")
            except Exception:
                pass
                
        if result["latest_epoch"] is None:
            latest_infer_path = os.path.join(run_path, "inference", "latest", "infer_config.json")
            if os.path.exists(latest_infer_path):
                try:
                    with open(latest_infer_path, "r", encoding="utf-8") as f:
                        cfg = json.load(f)
                    result["latest_epoch"] = cfg.get("epoch")
                except Exception:
                    pass

    return result


def classify_run(args, run_name):
    """
    Classify run into categories.
    """
    single_scene_only = args.get("single_scene_only", False)
    num_scenes = args.get("num_scenes", None)

    # Explicit single_scene_only flag
    if single_scene_only:
        return "single_scene"

    # Check num_scenes
    if num_scenes is not None:
        if num_scenes == 1:
            return "single_scene"
        elif num_scenes == 10:
            return "10scene"
        elif num_scenes == -1 or num_scenes == 0 or num_scenes >= 89:
            return "90scene"
        else:
            return f"other_{num_scenes}scene"

    # Fallback: infer from name
    if "scene0" in run_name:
        return "single_scene"

    return "unknown"


def get_inference_tags(run_path):
    """List the inference sub-tags (best, best_psnr, latest, etc.)."""
    infer_dir = os.path.join(run_path, "inference")
    if not os.path.isdir(infer_dir):
        return ""
    try:
        subdirs = sorted([
            d for d in os.listdir(infer_dir)
            if os.path.isdir(os.path.join(infer_dir, d))
        ])
        return ", ".join(subdirs)
    except Exception:
        return ""


def main():
    print(f"Scanning: {RUNS_DIR}")

    if not os.path.isdir(RUNS_DIR):
        print(f"ERROR: {RUNS_DIR} not found!")
        sys.exit(1)

    # Get all run directories
    all_dirs = sorted([
        d for d in os.listdir(RUNS_DIR)
        if os.path.isdir(os.path.join(RUNS_DIR, d)) and d.startswith("gm_")
    ])

    print(f"Total run directories: {len(all_dirs)}")

    valid_runs = []
    invalid_runs = []

    for run_name in all_dirs:
        run_path = os.path.join(RUNS_DIR, run_name)
        if os.path.isdir(os.path.join(run_path, "inference")):
            valid_runs.append(run_name)
        else:
            invalid_runs.append(run_name)

    print(f"Valid runs (with inference): {len(valid_runs)}")
    print(f"Invalid runs (no inference): {len(invalid_runs)}")

    # Collect all data
    all_data = []
    all_args_keys = set()

    for run_name in valid_runs:
        run_path = os.path.join(RUNS_DIR, run_name)
        args = load_args(run_path)
        all_args_keys.update(args.keys())

        psnr_info = extract_psnr_info(run_path)
        category = classify_run(args, run_name)
        infer_tags = get_inference_tags(run_path)

        record = {
            "run_name": run_name,
            "category": category,
            "avg_psnr": psnr_info["avg_psnr"],
            "best_epoch": psnr_info["best_epoch"],
            "latest_epoch": psnr_info["latest_epoch"],
            "plane0_psnr": psnr_info["plane0_psnr"],
            "plane20_psnr": psnr_info["plane20_psnr"],
            "plane39_psnr": psnr_info["plane39_psnr"],
            "ckpt_tag": psnr_info["ckpt_tag"],
            "inference_tags": infer_tags,
            "args": args,
        }
        all_data.append(record)

        psnr_str = f"{psnr_info['avg_psnr']:.2f}" if psnr_info['avg_psnr'] else "N/A"
        print(f"  [{category:>15}] {run_name} | avg_PSNR: {psnr_str} | epoch: {psnr_info['best_epoch']}")

    # Sort args keys - priority keys first
    priority_keys = [
        "arch", "diopter_mode", "energy_head", "channels", "bottleneck_modes",
        "use_film", "long_skip", "activation", "num_attn_heads",
        "epochs", "batch_size", "lr", "lr_scheduler", "weight_decay",
        "gm_steps", "gm_step_size", "eta_schedule", "eta_min",
        "langevin_noise", "noise_method", "noise_scale",
        "clip_image", "amp",
        "compositional_ebm", "enable_struct", "enable_percep", "enable_phys",
        "w_struct", "w_percep", "w_phys",
        "alpha_struct", "beta_struct",
        "enable_phys_blur", "enable_phys_occ", "enable_phys_energy", "enable_phys_bokeh",
        "lambda_struct", "lambda_percep", "lambda_phys",
        "lambda_blur_edge", "lambda_occlusion", "lambda_energy_conserv", "lambda_bokeh",
        "enable_energy_dist", "weight_energy_dist", "energy_dist_scale",
        "enable_energy_anchor", "weight_energy_anchor",
        "single_scene_only", "num_scenes",
        "train_bypass", "bypass_lambda", "bypass_gamma", "bypass_warmup", "bypass_ramp",
        "sharp_prior", "sharp_lambda", "sharp_gamma",
        "infer_sharp", "infer_sharp_lambda", "infer_sharp_gamma", "infer_sharp_start",
        "compile", "compile_mode",
        "interleave_rate", "phys_gamma", "kappa_occ", "energy_pool_k",
        "bokeh_threshold", "bokeh_dilate_k",
        "force_compositional", "new_run_on_resume",
        "unmatch_ratio", "save_every",
        "scene_idx", "plane_idx",
        "resume", "resume_dir",
        "data_dir", "generated_data_dir", "output_dir", "run_dir",
        "num_workers",
        "timestamp",
    ]

    sorted_args_keys = []
    remaining = set(all_args_keys)
    for k in priority_keys:
        if k in remaining:
            sorted_args_keys.append(k)
            remaining.discard(k)
    sorted_args_keys.extend(sorted(remaining))

    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Group by category
    categories = {}
    for record in all_data:
        cat = record["category"]
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(record)

    print(f"\n{'='*60}")
    print("Category breakdown:")
    for cat, records in sorted(categories.items()):
        print(f"  {cat}: {len(records)} runs")

    # Build CSV columns
    base_columns = [
        "run_name", "category",
        "avg_psnr", "best_epoch", "latest_epoch", "ckpt_tag",
        "plane0_psnr", "plane20_psnr", "plane39_psnr",
        "inference_tags",
    ]
    args_columns = [f"arg_{k}" for k in sorted_args_keys]
    all_columns = base_columns + args_columns

    def write_csv(filepath, records):
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=all_columns)
            writer.writeheader()
            for record in sorted(records, key=lambda r: -(r["avg_psnr"] or 0)):
                row = {}
                for col in base_columns:
                    row[col] = record.get(col, "")
                for k in sorted_args_keys:
                    val = record["args"].get(k, "")
                    # Convert booleans/None for cleaner display
                    if val is True:
                        val = "TRUE"
                    elif val is False:
                        val = "FALSE"
                    elif val is None:
                        val = ""
                    row[f"arg_{k}"] = val
                writer.writerow(row)
        print(f"  Written: {filepath} ({len(records)} runs)")

    # Write per-category CSVs
    for cat, records in sorted(categories.items()):
        csv_path = os.path.join(OUTPUT_DIR, f"runs_{cat}.csv")
        write_csv(csv_path, records)

    # Write combined CSV
    csv_all_path = os.path.join(OUTPUT_DIR, "runs_all.csv")
    write_csv(csv_all_path, all_data)

    # Write invalid runs list
    invalid_path = os.path.join(OUTPUT_DIR, "invalid_runs.txt")
    with open(invalid_path, "w", encoding="utf-8") as f:
        f.write(f"# Invalid runs (no inference folder): {len(invalid_runs)}\n\n")
        for name in invalid_runs:
            f.write(f"{name}\n")
    print(f"  Written: {invalid_path} ({len(invalid_runs)} runs)")

    # --- Excel output ---
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        xlsx_path = os.path.join(OUTPUT_DIR, "runs_summary.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        top_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        alt_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='999999'),
            right=Side(style='thin', color='999999'),
            top=Side(style='thin', color='999999'),
            bottom=Side(style='thin', color='999999'),
        )

        # Per-category sheets + ALL
        sheet_data = list(sorted(categories.items()))
        sheet_data.append(("ALL", all_data))

        for cat, records in sheet_data:
            sheet_name = cat[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Header
            for col_idx, col_name in enumerate(all_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = thin_border

            sorted_records = sorted(records, key=lambda r: -(r["avg_psnr"] or 0))

            for row_idx, record in enumerate(sorted_records, 2):
                for col_idx, col_name in enumerate(all_columns, 1):
                    if col_name in base_columns:
                        val = record.get(col_name, "")
                    else:
                        k = col_name[4:]  # strip "arg_"
                        val = record["args"].get(k, "")
                        if val is True:
                            val = "TRUE"
                        elif val is False:
                            val = "FALSE"
                        elif val is None:
                            val = ""

                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border

                    # Format numbers
                    if isinstance(val, float):
                        cell.number_format = '0.00'

                # Alternating row colors
                fill = None
                if row_idx == 2:
                    fill = top_fill  # Highlight top PSNR
                elif row_idx % 2 == 0:
                    fill = alt_fill

                if fill:
                    for col_idx in range(1, len(all_columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

            # Auto-fit column widths
            for col_idx in range(1, len(all_columns) + 1):
                max_len = len(str(all_columns[col_idx - 1]))
                for row_idx in range(2, min(len(sorted_records) + 2, 20)):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None:
                        max_len = max(max_len, min(len(str(cell_val)), 60))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(all_columns))}1"

        wb.save(xlsx_path)
        print(f"  Written: {xlsx_path}")

    except ImportError:
        print("  [NOTE] openpyxl not available. Install with: pip install openpyxl")
        print("         CSV files are still available.")

    print(f"\n{'='*60}")
    print(f"DONE! Results in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
